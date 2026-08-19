"""
Импорт справочника автомобилей из файла, скачанного вами с площадки.

Скрипт ничего не выкачивает из интернета — он читает файл, который вы
получили легально: справочник Авито Автозагрузки, выгрузку Дрома,
купленную базу или собственный CSV.

Запуск:
    python -m app.scripts.import_catalog --file spravochnik.xlsx --source avito
    python -m app.scripts.import_catalog --file catalog.csv --source drom --dry-run

Идемпотентен: повторный запуск того же файла ничего не дублирует,
только дополняет отсутствующее.
"""

import argparse
import asyncio
import csv
import re
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

from sqlalchemy import text

from ..database import SessionFactory

# ------------------------------------------------------------------
# Распознавание колонок: заголовки у всех источников разные
# ------------------------------------------------------------------

COLUMN_ALIASES = {
    "brand": ["марка", "марка авто", "mark", "make", "brand", "производитель"],
    "model": ["модель", "model"],
    "generation": ["поколение", "generation", "кузов", "серия", "конфигурация"],
    "body": ["тип кузова", "body", "body_type", "кузов тип"],
    "year_from": ["год начала", "год от", "year_from", "начало выпуска", "с года"],
    "year_to": ["год окончания", "год до", "year_to", "конец выпуска", "по год"],
    "years": ["годы", "годы выпуска", "years", "период"],
}

TRANSLIT = str.maketrans({
    "а":"a","б":"b","в":"v","г":"g","д":"d","е":"e","ё":"e","ж":"zh","з":"z",
    "и":"i","й":"y","к":"k","л":"l","м":"m","н":"n","о":"o","п":"p","р":"r",
    "с":"s","т":"t","у":"u","ф":"f","х":"h","ц":"c","ч":"ch","ш":"sh","щ":"sch",
    "ъ":"","ы":"y","ь":"","э":"e","ю":"yu","я":"ya",
})


def slugify(value: str) -> str:
    s = value.strip().lower().translate(TRANSLIT)
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "x"


def match_columns(headers: list[str]) -> dict[str, int]:
    """Сопоставить заголовки файла с нашими полями."""
    found = {}
    for idx, raw in enumerate(headers):
        key = (raw or "").strip().lower()
        for field, aliases in COLUMN_ALIASES.items():
            if key in aliases and field not in found:
                found[field] = idx
    return found


YEAR_RE = re.compile(r"(19|20)\d{2}")


def parse_years(value: str) -> tuple[int | None, int | None]:
    """«2011–2015», «с 2018», «2017 - н.в.» -> (2011, 2015) / (2018, None)"""
    if not value:
        return None, None
    years = [int(m.group()) for m in YEAR_RE.finditer(str(value))]
    if not years:
        return None, None
    if len(years) == 1:
        # «с 2018» или «2018 — н.в.»
        return years[0], None
    return years[0], years[1]


# ------------------------------------------------------------------
# Чтение файлов
# ------------------------------------------------------------------

def read_rows(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return read_xlsx(path)
    if suffix in (".csv", ".tsv", ".txt"):
        return read_csv(path)
    if suffix == ".xml":
        return read_xml(path)
    raise SystemExit(f"Не понимаю формат {suffix}. Нужен xlsx, csv или xml.")


def read_xlsx(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("Нужен openpyxl: pip install openpyxl")

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb.active
    rows = sheet.iter_rows(values_only=True)

    headers = [str(c) if c is not None else "" for c in next(rows)]
    cols = match_columns(headers)
    if "brand" not in cols or "model" not in cols:
        raise SystemExit(f"Не нашёл колонки марки и модели. Заголовки: {headers}")

    out = []
    for row in rows:
        out.append(extract(row, cols))
    wb.close()
    return out


def read_csv(path: Path) -> list[dict]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        reader = csv.reader(f, dialect)
        headers = next(reader)
        cols = match_columns(headers)
        if "brand" not in cols or "model" not in cols:
            raise SystemExit(f"Не нашёл колонки марки и модели. Заголовки: {headers}")
        return [extract(row, cols) for row in reader]


def read_xml(path: Path) -> list[dict]:
    """Плоская структура вида <Mark name=..><Model name=..><Generation name=../>"""
    tree = ET.parse(path)
    out = []

    def attr(node, *names):
        for n in names:
            v = node.get(n) or node.get(n.capitalize())
            if v:
                return v
        return None

    for mark in tree.iter():
        if mark.tag.lower() not in ("mark", "make", "brand", "марка"):
            continue
        brand = attr(mark, "name", "id") or (mark.text or "").strip()
        for model in mark:
            model_name = attr(model, "name") or (model.text or "").strip()
            gens = list(model)
            if not gens:
                out.append({"brand": brand, "model": model_name})
                continue
            for gen in gens:
                y_from, y_to = parse_years(attr(gen, "years", "year") or "")
                out.append({
                    "brand": brand, "model": model_name,
                    "generation": attr(gen, "name"),
                    "body": attr(gen, "body", "bodytype"),
                    "year_from": y_from, "year_to": y_to,
                })
    return out


def extract(row, cols: dict) -> dict:
    def cell(field):
        idx = cols.get(field)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if v is not None and str(v).strip() else None

    y_from, y_to = None, None
    if "years" in cols:
        y_from, y_to = parse_years(cell("years") or "")
    if cell("year_from"):
        y_from, _ = parse_years(cell("year_from"))
    if cell("year_to"):
        y_to, _ = parse_years(cell("year_to"))

    return {
        "brand": cell("brand"),
        "model": cell("model"),
        "generation": cell("generation"),
        "body": cell("body"),
        "year_from": y_from,
        "year_to": y_to,
    }


# ------------------------------------------------------------------
# Загрузка в базу
# ------------------------------------------------------------------

async def load(rows: list[dict], source: str, filename: str, dry_run: bool):
    stats = {"brands": 0, "models": 0, "generations": 0, "skipped": 0, "total": len(rows)}
    brand_cache: dict[str, int] = {}
    model_cache: dict[tuple[int, str], int] = {}

    async with SessionFactory() as session:
        for row in rows:
            if not row.get("brand") or not row.get("model"):
                stats["skipped"] += 1
                continue

            brand_name = row["brand"]
            if brand_name not in brand_cache:
                bid = await upsert_brand(session, brand_name, source, stats)
                brand_cache[brand_name] = bid
            brand_id = brand_cache[brand_name]

            key = (brand_id, row["model"])
            if key not in model_cache:
                mid = await upsert_model(session, brand_id, row["model"], source, stats)
                model_cache[key] = mid
            model_id = model_cache[key]

            if row.get("generation"):
                await upsert_generation(session, model_id, row, source, stats)

        if dry_run:
            await session.rollback()
            print("\n[ПРОБНЫЙ ЗАПУСК] Изменения откачены, ничего не сохранено.")
        else:
            await session.execute(text("""
                INSERT INTO import_log (source, filename, brands_new, models_new,
                                        generations_new, rows_total, rows_skipped,
                                        finished_at)
                VALUES (:s, :f, :b, :m, :g, :t, :sk, now())
            """), {"s": source, "f": filename, "b": stats["brands"],
                   "m": stats["models"], "g": stats["generations"],
                   "t": stats["total"], "sk": stats["skipped"]})
            await session.commit()

    return stats


async def upsert_brand(session, name: str, source: str, stats) -> int:
    slug = slugify(name)
    row = (await session.execute(text("""
        INSERT INTO brands (name, slug, avito_name, source)
        VALUES (:n, :s, CASE WHEN :src = 'avito' THEN :n END, :src)
        ON CONFLICT (slug) DO UPDATE
            SET avito_name = COALESCE(brands.avito_name,
                    CASE WHEN :src = 'avito' THEN :n END)
        RETURNING id, (xmax = 0) AS inserted
    """), {"n": name, "s": slug, "src": source})).first()
    if row.inserted:
        stats["brands"] += 1
    return row.id


async def upsert_model(session, brand_id: int, name: str, source: str, stats) -> int:
    slug = slugify(name)
    row = (await session.execute(text("""
        INSERT INTO models (brand_id, name, slug, avito_name, source)
        VALUES (:b, :n, :s, CASE WHEN :src = 'avito' THEN :n END, :src)
        ON CONFLICT (brand_id, slug) DO UPDATE
            SET avito_name = COALESCE(models.avito_name,
                    CASE WHEN :src = 'avito' THEN :n END)
        RETURNING id, (xmax = 0) AS inserted
    """), {"b": brand_id, "n": name, "s": slug, "src": source})).first()
    if row.inserted:
        stats["models"] += 1
    return row.id


async def upsert_generation(session, model_id: int, row: dict, source: str, stats):
    # Год начала обязателен по схеме. Нет года — ставим заглушку
    # и помечаем на проверку, чтобы строка не потерялась.
    y_from = row.get("year_from")
    needs_review = y_from is None
    result = (await session.execute(text("""
        INSERT INTO generations (model_id, name, body_type, year_from, year_to,
                                 avito_name, source, needs_review)
        VALUES (:m, :n, :b, COALESCE(:yf, 1900), :yt,
                CASE WHEN :src = 'avito' THEN :n END, :src, :rev)
        ON CONFLICT (model_id, name) DO UPDATE
            SET body_type = COALESCE(generations.body_type, EXCLUDED.body_type),
                year_to   = COALESCE(generations.year_to, EXCLUDED.year_to),
                avito_name = COALESCE(generations.avito_name, EXCLUDED.avito_name)
        RETURNING (xmax = 0) AS inserted
    """), {"m": model_id, "n": row["generation"], "b": row.get("body"),
           "yf": y_from, "yt": row.get("year_to"), "src": source,
           "rev": needs_review})).first()
    if result.inserted:
        stats["generations"] += 1


# ------------------------------------------------------------------
# CLI
# ------------------------------------------------------------------

async def main():
    ap = argparse.ArgumentParser(description="Импорт справочника автомобилей")
    ap.add_argument("--file", required=True, help="xlsx, csv или xml со справочником")
    ap.add_argument("--source", default="manual",
                    choices=["avito", "drom", "wikidata", "manual"],
                    help="откуда файл — влияет на заполнение названий для фидов")
    ap.add_argument("--dry-run", action="store_true",
                    help="разобрать файл и показать итог, ничего не записывая")
    ap.add_argument("--limit", type=int, help="обработать только первые N строк")
    args = ap.parse_args()

    path = Path(args.file)
    if not path.exists():
        sys.exit(f"Файл не найден: {path}")

    print(f"Читаю {path.name}…")
    rows = read_rows(path)
    if args.limit:
        rows = rows[: args.limit]
    print(f"Строк в файле: {len(rows)}")

    if rows[:1]:
        print(f"Первая строка: {rows[0]}")

    stats = await load(rows, args.source, path.name, args.dry_run)

    print(f"""
Готово.
  Обработано строк:   {stats['total']}
  Пропущено (пусто):  {stats['skipped']}
  Новых марок:        {stats['brands']}
  Новых моделей:      {stats['models']}
  Новых поколений:    {stats['generations']}
""")
    if stats["generations"]:
        print("Поколения без года выпуска помечены needs_review — "
              "проверьте их в разделе справочника.")


if __name__ == "__main__":
    asyncio.run(main())
